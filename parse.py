import openpyxl as excel
import re


def parse(fn):
    wb = excel.load_workbook(filename=fn)
    ws = wb.active

    dep = [cell.value for cell in ws['C'] if cell.value]
    scr = [cell.value for cell in ws['D'] if cell.value]

    data = zip(dep, scr)

    a1 = re.compile('학과')
    a2 = re.compile('학부')

    b1 = re.compile('학')
    b2 = re.compile('과')

    c1 = re.compile('연대')
    c2 = re.compile('연세대')
    c3 = re.compile('연')

    d = re.compile('\(.*\)')

    rex = [a1, a2, b1, b2, c1, c2, c3, d]

    e = re.compile('\d*')

    dep = []
    stopwords = [
        ['정치외교', '정외'],
        ['컴퓨터공', '컴', '컴퓨터과', '컴퓨터'],
        ['HASS', '하스', '융합인문사회계열', 'Hass', 'hass', '융합인문사회과'],
        ['독어독문', '독문'],
        ['학교수', '수학교육'],
        ['화생공', '화공생명공', '화공생명과', '화공생명'],
        ['전전공', '전기전자공', '전자전기공', '전기전자'],
        ['수', '교수'],
        ['사환시', '사회환경시스템', '사회환경시스템공'],
        ['시반공', '시스템반도체', '시스템반도체공'],
        ['신소재공', '신소재']

    ]
    lines =[
        ['(자연)', '(자연계열)'],
        ['(인문)', '(인문계열)']
    ]

    sheet = []

    for cell in data:
        line = None
        cell = list(cell)

        cell[0] = cell[0].replace(' ', '')

        if d.search(cell[0]):
            line = d.findall(cell[0])[0]

        for r in rex:
            cell[0] = r.sub('', cell[0], count=1)

        for stop in stopwords:
            if cell[0] in stop:
                cell[0] = stop[0]

        if line:
            for l in lines:
                if line in l:
                    cell[0] = cell[0]+l[0]

        dep.append(cell[0])

        cell[1] = float(cell[1])
        sheet.append(cell)

    dep = list(set(dep))
    database = dict()
    for d in dep:
        database[d] = []

    for data in sheet:
        database[data[0]].append(data[1])

    wwb = excel.Workbook()
    wws = wwb.active
    wws.title = '연세대 점공'

    for i, (key, val) in enumerate(sorted(database.items())):
        database[key] = sorted(val)
        wws.append(
            [key]+val
        )

    wwb.save(filename='연세대학교 점수공개 결과.xlsx')
    wwb.close()
